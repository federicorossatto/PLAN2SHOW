"""Excel workbook creation and reading for PLAN2SHOW tracks."""

import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from constants import APP_AUTHOR, APP_NAME, REQUIRED_COLUMNS

INVALID_FILENAME_CHARS_RE = re.compile(r'[\\/*?:"<>|]')
def cell_to_text(value):
    """Convert an Excel cell value to a clean compiler string."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()

def read_excel_rows(excel_path):
    """Read cue rows and the track title from a PLAN2SHOW workbook."""
    if not excel_path.lower().endswith(".xlsx"):
        raise ValueError("PLAN2SHOW expects an .xlsx track file.")

    try:
        workbook = load_workbook(excel_path, data_only=True, read_only=True)
    except (OSError, ValueError) as error:
        raise ValueError(
            f"Could not open Excel file '{os.path.basename(excel_path)}': {error}"
        ) from error

    try:
        if "TIMELINE" not in workbook.sheetnames:
            raise ValueError("The workbook does not contain the required 'TIMELINE' sheet.")

        worksheet = workbook["TIMELINE"]
        track_title = cell_to_text(worksheet["B2"].value) or None
        header_row = None
        fieldnames = []

        for row_number in range(1, min(worksheet.max_row, 30) + 1):
            values = [
                cell_to_text(worksheet.cell(row=row_number, column=column).value)
                for column in range(1, len(REQUIRED_COLUMNS) + 1)
            ]
            if "Cue Number" in values and "Actions" in values:
                header_row = row_number
                fieldnames = values
                break

        if header_row is None:
            raise ValueError(
                "Could not find the PLAN2SHOW headers in TIMELINE. "
                f"Expected: {', '.join(REQUIRED_COLUMNS)}"
            )

        missing = [column for column in REQUIRED_COLUMNS if column not in fieldnames]
        if missing:
            raise ValueError("Missing required Excel columns: " + ", ".join(missing))

        rows = []
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            values = [
                cell_to_text(worksheet.cell(row=row_number, column=column).value)
                for column in range(1, len(fieldnames) + 1)
            ]
            row = dict(zip(fieldnames, values))
            if not any(row.get(column, "") for column in REQUIRED_COLUMNS):
                continue
            row["_source_row"] = str(row_number)
            rows.append(row)

        if not rows:
            raise ValueError("The TIMELINE sheet does not contain any cue rows to compile.")

        return rows, track_title
    finally:
        workbook.close()

def sanitize_component(raw_text, fallback="Untitled"):
    """Return a filesystem-safe component."""
    text = (raw_text or "").strip() or fallback
    return INVALID_FILENAME_CHARS_RE.sub("_", text).replace(" ", "_")


def create_excel_template(tracks_directory):
    """
    Create a styled Excel workbook for a PLAN2SHOW track.

    The workbook contains:
    - a TIMELINE sheet for cue programming;
    - a HELP sheet with Actions syntax examples;
    - validation rules for numeric fields;
    - formatted input rows.
    """
    print("\n--- Create Excel Track Template ---")

    raw_name = input("Track name: ").strip()

    display_name = raw_name or "New Show"
    safe_name = sanitize_component(
        raw_name,
        fallback="New_Show"
    )

    file_name = safe_name + ".xlsx"

    target_path = os.path.join(
        tracks_directory,
        file_name
    )

    if os.path.exists(target_path):
        overwrite = input(
            f"[WARNING] '{file_name}' already exists. "
            "Overwrite? (y/n): "
        ).strip().lower()

        if overwrite != "y":
            print("[INFO] Excel template creation cancelled.")
            return

    workbook = Workbook()

    timeline = workbook.active
    timeline.title = "TIMELINE"

    help_sheet = workbook.create_sheet("HELP")

    # ------------------------------------------------------------------
    # Color palette
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # Visual theme
    # ------------------------------------------------------------------

    navy = "14213D"
    dark_blue = "1F4E78"
    header_blue = "2F75B5"
    accent_blue = "5B9BD5"

    title_white = "FFFFFF"
    subtitle_blue = "D9EAF7"

    row_white = "FFFFFF"
    row_blue = "EAF3F8"
    notes_white = "FFFDF2"
    notes_yellow = "FFF2CC"

    help_green = "E2F0D9"
    error_red = "FFC7CE"

    text_gray = "59636E"
    input_text_blue = "0070C0"

    border_light = "B4C7DC"
    border_medium = "7F9DB9"
    border_dark = "1F4E78"

    white = title_white

    thin_side = Side(
    style="thin",
    color=border_light
    )

    medium_side = Side(
        style="medium",
        color=border_medium
    )

    dark_side = Side(
        style="medium",
        color=border_dark
    )

    data_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    header_border = Border(
        left=medium_side,
        right=medium_side,
        top=dark_side,
        bottom=dark_side
    )

    track_border = Border(
        left=dark_side,
        right=dark_side,
        top=dark_side,
        bottom=dark_side
    )
    # ------------------------------------------------------------------
    # TIMELINE title
    # ------------------------------------------------------------------

    timeline.merge_cells("A1:H1")
    timeline["A1"] = "PLAN2SHOW  |  LIGHTING TIMELINE  |  by Federico Rossatto"

    timeline["A1"].font = Font(
        name="Calibri",
        size=22,
        bold=True,
        color=title_white
    )

    timeline["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor=navy
    )

    timeline["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    timeline["A1"].border = Border(
        bottom=Side(
            style="thick",
            color=accent_blue
        )
    )

    timeline.row_dimensions[1].height = 42

    # ------------------------------------------------------------------
    # Track information
    # ------------------------------------------------------------------

    timeline["A2"] = "TRACK"

    timeline["A2"].font = Font(
        name="Calibri",
        size=11,
        bold=True,
        color=title_white
    )

    timeline["A2"].fill = PatternFill(
        fill_type="solid",
        fgColor=dark_blue
    )

    timeline["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    timeline["A2"].border = track_border

    timeline.merge_cells("B2:H2")
    timeline["B2"] = display_name

    timeline["B2"].font = Font(
        name="Calibri",
        size=14,
        bold=True,
        color=input_text_blue
    )

    timeline["B2"].fill = PatternFill(
        fill_type="solid",
        fgColor=subtitle_blue
    )

    timeline["B2"].alignment = Alignment(
        horizontal="left",
        vertical="center",
        indent=1
    )

    timeline["B2"].border = track_border
    timeline.row_dimensions[2].height = 30

    # ------------------------------------------------------------------
    # Instructions
    # ------------------------------------------------------------------

    timeline.merge_cells("A3:H3")

    timeline["A3"] = (
        "Enter one cue per row. Minutes and Seconds are relative to "
        "the beginning of the track. Open the HELP sheet for the "
        "complete Actions syntax."
    )

    timeline["A3"].font = Font(
        name="Calibri",
        size=10,
        italic=True,
        color=text_gray
    )

    timeline["A3"].fill = PatternFill(
        fill_type="solid",
        fgColor="F3F6F9"
    )

    timeline["A3"].alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
        indent=1
    )

    timeline["A3"].border = Border(
        bottom=thin_side
    )

    timeline.row_dimensions[3].height = 30
    timeline.row_dimensions[4].height = 8

    # ------------------------------------------------------------------
    # Timeline headers
    # ------------------------------------------------------------------

    header_row = 5
    first_data_row = 6
    last_data_row = 205

    for column_number, column_name in enumerate(
        REQUIRED_COLUMNS,
        start=1
    ):
        cell = timeline.cell(
            row=header_row,
            column=column_number,
            value=column_name
        )

        cell.font = Font(
            name="Calibri",
            size=11,
            bold=True,
            color=title_white
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=header_blue
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = header_border

    timeline.row_dimensions[header_row].height = 32

    # ------------------------------------------------------------------
    # Input rows
    # ------------------------------------------------------------------

    for row_number in range(first_data_row, last_data_row + 1):
        is_even_row = row_number % 2 == 0

        standard_fill = PatternFill(
            fill_type="solid",
            fgColor=row_white if is_even_row else row_blue
        )
        notes_fill = PatternFill(
            fill_type="solid",
            fgColor=notes_white if is_even_row else notes_yellow
        )

        for column_number in range(1, 9):
            cell = timeline.cell(row=row_number, column=column_number)
            cell.fill = notes_fill if column_number == 8 else standard_fill
            cell.font = Font(name="Calibri", size=10, color=input_text_blue)
            cell.border = data_border

            if column_number in (1, 3, 4, 5, 6):
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
            else:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True
                )

        timeline.row_dimensions[row_number].height = 24

    # Apply explicit numeric formats.
    for row_number in range(first_data_row, last_data_row + 1):
        timeline.cell(row=row_number, column=1).number_format = "0.##"
        timeline.cell(row=row_number, column=3).number_format = "0"
        timeline.cell(row=row_number, column=4).number_format = "0"
        timeline.cell(row=row_number, column=5).number_format = "0.##"
        timeline.cell(row=row_number, column=6).number_format = "0"

    # ------------------------------------------------------------------
    # First input row
    # ------------------------------------------------------------------

    # Keep the first row empty so the template never contains a demo cue
    # that could be compiled accidentally. Usage examples are available
    # in the HELP sheet.

    # ------------------------------------------------------------------
    # Data validation
    # ------------------------------------------------------------------

    cue_validation = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=True
    )

    minute_validation = DataValidation(
    type="whole",
    operator="greaterThanOrEqual",
    formula1="0",
    allow_blank=True
    )

    second_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="59",
        allow_blank=True
    )

    fade_validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True
    )

    bpm_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="999",
        allow_blank=True
    )

    cue_validation.errorTitle = "Invalid Cue"
    cue_validation.error = (
        "Cue must be a positive number."
    )
    cue_validation.showErrorMessage = True

    minute_validation.errorTitle = "Invalid Minutes"
    minute_validation.error = (
    "Minutes must be a whole number greater than or equal to zero."
    )
    minute_validation.showErrorMessage = True

    second_validation.errorTitle = "Invalid Seconds"
    second_validation.error = (
        "Seconds must be between 0 and 59."
    )
    second_validation.showErrorMessage = True

    fade_validation.errorTitle = "Invalid Fade"
    fade_validation.error = (
        "Fade must be zero or a positive number."
    )
    fade_validation.showErrorMessage = True

    bpm_validation.errorTitle = "Invalid BPM"
    bpm_validation.error = (
        "BPM must be a whole number between 1 and 999."
    )
    bpm_validation.showErrorMessage = True

    for validation in (
        cue_validation,
        minute_validation,
        second_validation,
        fade_validation,
        bpm_validation
    ):
        timeline.add_data_validation(validation)

    cue_validation.add(
        f"A{first_data_row}:A{last_data_row}"
    )

    minute_validation.add(
        f"C{first_data_row}:C{last_data_row}"
    )

    second_validation.add(
        f"D{first_data_row}:D{last_data_row}"
    )

    fade_validation.add(
        f"E{first_data_row}:E{last_data_row}"
    )

    bpm_validation.add(
        f"F{first_data_row}:F{last_data_row}"
    )

    # ------------------------------------------------------------------
    # Conditional formatting
    # ------------------------------------------------------------------

    timeline.conditional_formatting.add(
    f"D{first_data_row}:D{last_data_row}",
    CellIsRule(
        operator="greaterThan",
        formula=["59"],
        fill=PatternFill(
            fill_type="solid",
            fgColor=error_red
        )
    )
)

    timeline.conditional_formatting.add(
        f"E{first_data_row}:E{last_data_row}",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=PatternFill(
                fill_type="solid",
                fgColor=error_red
            )
        )
    )

    # ------------------------------------------------------------------
    # Filters, dimensions and view
    # ------------------------------------------------------------------

    timeline.auto_filter.ref = (
        f"A{header_row}:H{last_data_row}"
    )

    column_widths = {
        "A": 16,
        "B": 24,
        "C": 11,
        "D": 11,
        "E": 10,
        "F": 10,
        "G": 58,
        "H": 34
    }

    for column_letter, width in column_widths.items():
        timeline.column_dimensions[
            column_letter
        ].width = width

    timeline.freeze_panes = "A6"
    timeline.sheet_view.showGridLines = False
    timeline.sheet_view.zoomScale = 90
    timeline.sheet_view.zoomScaleNormal = 90
    timeline.sheet_view.selection[0].activeCell = "A6"
    timeline.sheet_view.selection[0].sqref = "A6"

    timeline.page_setup.orientation = "landscape"
    timeline.page_setup.fitToWidth = 1
    timeline.page_setup.fitToHeight = 0
    timeline.print_title_rows = "1:5"

    # ------------------------------------------------------------------
    # HELP sheet
    # ------------------------------------------------------------------

    help_sheet.merge_cells("A1:D1")
    help_sheet["A1"] = "PLAN2SHOW ACTIONS GUIDE"

    help_sheet["A1"].font = Font(
        name="Calibri",
        size=18,
        bold=True,
        color=white
    )

    help_sheet["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor=dark_blue
    )

    help_sheet["A1"].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    help_sheet.row_dimensions[1].height = 32

    help_headers = [
        "Syntax",
        "Meaning",
        "Example",
        "Result"
    ]

    for column_number, header in enumerate(
        help_headers,
        start=1
    ):
        cell = help_sheet.cell(
            row=3,
            column=column_number,
            value=header
        )

        cell.font = Font(
            bold=True,
            color=white
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=header_blue
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    help_rows = [
        [
            ":",
            "Separates the subject from its actions",
            "Sides: Blue",
            "Apply Blue to Sides"
        ],
        [
            ",",
            "Separates multiple actions",
            "Sides: Blue, 30%",
            "Apply Blue and intensity 30%"
        ],
        [
            ";",
            "Separates independent action blocks",
            "Sides: Blue ; Smoke: 25%",
            "Compile two subjects"
        ],
        [
            "..",
            "Defines a fixture range",
            "White 1..White 4: Full",
            "Select White 1 through White 4"
        ],
        [
            "+",
            "Combines fixtures",
            "Blinder 1+Blinder 2: Strobe",
            "Select both blinders"
        ]
    ]

    for row_number, row_values in enumerate(
        help_rows,
        start=4
    ):
        for column_number, value in enumerate(
            row_values,
            start=1
        ):
            cell = help_sheet.cell(
                row=row_number,
                column=column_number,
                value=value
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=help_green
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    help_sheet["A11"] = "Complete examples"

    help_sheet["A11"].font = Font(
        bold=True,
        color=white
    )

    help_sheet["A11"].fill = PatternFill(
        fill_type="solid",
        fgColor=header_blue
    )

    examples = [
        "Smoke: 25%",
        "Full Rig: Blackout",
        "Sides: Blue, 30%",
        "White 1..White 4: Full",
        "Blinder 1+Blinder 2: Strobe",
        "BPM Show: 120 BPM"
    ]

    for row_number, example in enumerate(
        examples,
        start=12
    ):
        help_sheet.cell(
            row=row_number,
            column=1,
            value=example
        )

    help_sheet.column_dimensions["A"].width = 36
    help_sheet.column_dimensions["B"].width = 42
    help_sheet.column_dimensions["C"].width = 45
    help_sheet.column_dimensions["D"].width = 45

    help_sheet.freeze_panes = "A4"
    help_sheet.sheet_view.showGridLines = False

    # ------------------------------------------------------------------
    # Save workbook
    # ------------------------------------------------------------------

    try:
        workbook.save(target_path)
    except OSError as error:
        raise ValueError(
            f"Could not save Excel template: {error}"
        ) from error

    print(f"[OK] Excel template created: {target_path}")
    print(
        "     Open the TIMELINE sheet in Excel and enter "
        "one cue per row."
    )

