"""Excel template and reader integration tests."""

import builtins

import pytest
from openpyxl import Workbook, load_workbook

from constants import REQUIRED_COLUMNS
from excel_io import create_excel_template, read_excel_rows


def test_create_template(monkeypatch, tmp_path):
    monkeypatch.setattr(builtins, "input", lambda _prompt="": "Dancing Queen")
    create_excel_template(str(tmp_path))
    path = tmp_path / "Dancing_Queen.xlsx"
    assert path.exists()

    workbook = load_workbook(path)
    assert workbook.sheetnames == ["TIMELINE", "HELP"]
    worksheet = workbook["TIMELINE"]
    assert worksheet["B2"].value == "Dancing Queen"
    assert [worksheet.cell(5, column).value for column in range(1, 9)] == REQUIRED_COLUMNS
    assert worksheet.column_dimensions["A"].width == 16
    workbook.close()


def test_read_template_rows(monkeypatch, tmp_path):
    monkeypatch.setattr(builtins, "input", lambda _prompt="": "Test Track")
    create_excel_template(str(tmp_path))
    path = tmp_path / "Test_Track.xlsx"
    workbook = load_workbook(path)
    worksheet = workbook["TIMELINE"]
    values = [1, "Intro", 0, 5, 2, 120, "Smoke: 25%", "Open"]
    for column, value in enumerate(values, start=1):
        worksheet.cell(6, column).value = value
    workbook.save(path)
    workbook.close()

    rows, title = read_excel_rows(str(path))
    assert title == "Test Track"
    assert len(rows) == 1
    assert rows[0]["Cue Number"] == "1"
    assert rows[0]["Seconds"] == "5"
    assert rows[0]["_source_row"] == "6"


def test_missing_timeline_sheet_is_rejected(tmp_path):
    path = tmp_path / "bad.xlsx"
    workbook = Workbook()
    workbook.active.title = "OTHER"
    workbook.save(path)
    workbook.close()
    with pytest.raises(ValueError, match="TIMELINE"):
        read_excel_rows(str(path))


def test_empty_timeline_is_rejected(tmp_path):
    path = tmp_path / "empty.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TIMELINE"
    worksheet["B2"] = "Empty"
    for column, name in enumerate(REQUIRED_COLUMNS, start=1):
        worksheet.cell(5, column).value = name
    workbook.save(path)
    workbook.close()
    with pytest.raises(ValueError, match="does not contain any cue rows"):
        read_excel_rows(str(path))


def test_non_xlsx_path_is_rejected(tmp_path):
    with pytest.raises(ValueError, match=".xlsx"):
        read_excel_rows(str(tmp_path / "track.csv"))
